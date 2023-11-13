import os
import numpy as np
import pandas as pd
import pickle
from .Graph import Graph
from .Style import Style
from scipy.stats import gaussian_kde
import plotly.graph_objects as go
from openpyxl import Workbook, load_workbook

from typing import Union, Dict, Callable, List, Optional, Tuple, cast

Number = Union[float, int, np.float64]
Trace = Union[go.Histogram, go.Scatter]


def export_metrics_to_table(
    data_list: List[Dict[str, Number]],
    threshold: float,
    save_path: str,
    save_fname: str,
):
    rounder = lambda x: f"{x:.2f}"
    wb = Workbook()
    ws = wb.active
    cols = "B C D E F G H I J K L".split()
    row = 2
    headers = [
        "Iteration",
        f"Percent > {threshold}",
        "Q1",
        "Q2",
        "Mean",
        "Q3",
        "Max",
        "Std",
    ]
    for col, header in zip(cols, headers):
        ws[col + str(row)] = header
    row += 1
    for i, data in enumerate(data_list):
        data.update({"Iteration": i})
        for col, header in zip(cols, headers):
            ws[col + str(row)] = rounder(data[header])
        row += 1
    wb.save(save_path + save_fname + ".xlsx")


def create_combined_workbook(tables_path: str, save_path: str):
    combined_wb = Workbook()
    combined_ws = combined_wb.active
    combined_ws.title = "Combined Data"
    current_row = 2

    for file_name in os.listdir(tables_path):
        if file_name.endswith(".xlsx"):
            file_path = os.path.join(tables_path, file_name)
            wb = load_workbook(file_path)
            ws = wb["Sheet"]

            combined_ws.cell(row=current_row, column=2, value=file_name)
            current_row += 1

            for row in range(2, 9):
                for col in range(2, 10):
                    combined_ws.cell(
                        row=current_row,
                        column=col,
                        value=ws.cell(row=row, column=col).value,
                    )
                current_row += 1

            current_row += 1

    combined_wb.save(save_path + "combined_tables.xlsx")


def load_dist(scoring_path: str, fname: str, numpify: bool = True) -> pd.DataFrame:
    df = pd.read_csv(scoring_path + fname + ".csv")
    if numpify:
        return df["score"].to_numpy()
    return df


def compute_cluster_scores(
    scoring_path: str, fname: str, aggregation_mode: str, keep_cluster_id: bool = False
) -> Union[np.ndarray, Dict]:
    scored_mols = load_dist(scoring_path, fname, numpify=False)
    cluster_to_scores: Dict[int, List[Union[float, int]]] = {}
    for _, row in scored_mols.iterrows():
        cluster_to_scores.setdefault(row["cluster_id"], []).append(row["score"])
    match aggregation_mode:
        case "mean":
            aggregator: Callable = np.mean
        case "median":
            aggregator = np.median
        case _:
            raise ValueError(f"Unknown aggregation mode: {aggregation_mode}")
    if keep_cluster_id:
        return {
            cluster_id: aggregator(scores)
            for cluster_id, scores in cluster_to_scores.items()
        }
    else:
        return np.array(
            [aggregator(scores) for cluster_id, scores in cluster_to_scores.items()]
        )


# if __name__ == "__main__":
#     import secret
#     df = load_dist(
#         scoring_path=secret.PRODUCTION_RUNS_PATH+"5. Scoring/scored_dataframes/",
#         fname="model7_baseline_mix_k100",
#     )
#     dd = compute_cluster_scores(
#         scoring_path=secret.PRODUCTION_RUNS_PATH+"5. Scoring/scored_dataframes/",
#         fname="model7_baseline_mix_k100",
#         aggregation_mode="mean",
#     )
#     print(np.array(dd).shape)


def prepare_cycle_config(
    prefix: str,
    descriptors_type: str,
    n_iters: int,
    channel: str,
    n_clusters: Optional[int] = None,
    filters: Optional[str] = None,
    target: Optional[str] = None,
) -> Tuple[List[str], List[str]]:
    if filters is not None:
        # filtering runs
        assert isinstance(
            target, str
        )  # only because we started specifying targets for filtered runs
        assert filters in {"ADMET", "ADMET+FGs"}
        if "model2" in prefix:
            fnames = [
                f"{prefix[:6]}_baseline_{target.upper()}_{descriptors_type}_k{n_clusters}_{filters}",
                *(
                    f"{prefix}_al{i}_{channel}_{target.upper()}_{descriptors_type}_k{n_clusters}_{filters}"
                    for i in range(1, n_iters + 1)
                ),
            ]
        elif "model7" in prefix:
            fnames = [
                f"{prefix[:6]}_baseline_{target.upper()}_{descriptors_type}_k{n_clusters}_{filters}",
                *(
                    f"{prefix}_{channel}_al{i}_{target.upper()}_{descriptors_type}_k{n_clusters}_{filters}"
                    for i in range(1, n_iters + 1)
                ),
            ]
    elif n_clusters is not None:
        # original runs
        fnames = [
            f"{prefix}_baseline_{descriptors_type}_k{n_clusters}",
            *(
                f"{prefix}_{descriptors_type}{n_clusters}_{channel}_al{i}_{descriptors_type}_k{n_clusters}"
                for i in range(1, n_iters + 1)
            ),
        ]
    else:
        # random runs
        fnames = [
            f"{prefix}_baseline_{channel}",
            *(f"{prefix}_{channel}_al{i}_{channel}" for i in range(1, n_iters + 1)),
        ]
    labels = [f"Iteration {i}" for i in range(0, n_iters + 1)]
    return fnames, labels


def compute_metrics(data: np.ndarray, threshold: float) -> Dict[str, Number]:
    above_threshold_pct = np.sum(data >= threshold) / len(data) * 100
    q25, q50, q75 = np.percentile(data, [25, 50, 75], method="linear")
    return {
        f"Percent > {threshold}": above_threshold_pct,
        "Q1": q25,
        "Q2": q50,
        "Q3": q75,
        "Mean": data.mean(),
        "Max": data.max(),
        "Std": data.std(),
    }


def create_hist_trace(
    data: np.ndarray,
    label: str,
    color: str,
    bin_step: int,
    trace_opacity: float,
    density_line_opacity: float = 0.8,
    density_fill: Optional[str] = None,
    density_fill_opacity: float = 0.1,
    showlegend=True,
    visible=True,
) -> Tuple[go.Bar, go.Scatter]:
    density = gaussian_kde(data)
    xs = np.linspace(np.min(data), np.max(data), 200)

    hist_vals, bin_edges = np.histogram(
        data, bins=range(0, int(np.max(data)) + 2, bin_step), density=True
    )
    hist_trace = go.Bar(
        x=bin_edges[:-1],
        y=hist_vals,
        name=label,
        opacity=trace_opacity,
        marker=dict(
            color=Style.biscale[color][1],
            line=dict(color=Style.biscale[color][0](density_line_opacity), width=2),
        ),
        hovertemplate=[f"[{int(i)}, {int(i + bin_step)})" for i in bin_edges[:-1]],
        showlegend=showlegend,
    )

    density_trace = go.Scatter(
        x=xs,
        y=density(xs),
        mode="lines",
        name=label,
        line=dict(color=Style.biscale[color][0](density_line_opacity), width=4),
        fill=density_fill,
        fillcolor=Style.biscale[color][0](density_fill_opacity),
        showlegend=showlegend,
        visible=visible,
    )
    return hist_trace, density_trace


def prepare_score_distribution_traces(
    scoring_path: str,
    fnames: List[str],
    labels: List[str],
    colors: List[str],
    clusterize: bool = False,
    aggregation_mode: Optional[str] = None,
    threshold: float = 11,
    bin_step: int = 2,
    trace_opacity: float = 0.6,
    show_bars: bool = True,
    show_density: bool = True,
    density_line_opacity: float = 0.8,
    density_fill: Optional[str] = None,
    density_fill_opacity: float = 0.1,
    showlegend: bool = True,
) -> Tuple[List[Trace], List[Dict[str, Number]]]:
    traces, metrics_list = [], []
    if clusterize:
        assert aggregation_mode in {"mean", "median"}
        loader = lambda fname: compute_cluster_scores(scoring_path, fname, aggregation_mode)
    else:
        loader = lambda fname: load_dist(scoring_path, fname)
    for i, (fname, label, color) in enumerate(zip(fnames, labels, colors)):
        data = cast(np.ndarray, loader(fname))
        hist_trace, density_trace = create_hist_trace(
            data=data,
            label=label,
            color=color,
            bin_step=bin_step,
            trace_opacity=trace_opacity,
            density_line_opacity=density_line_opacity,
            density_fill=density_fill,
            density_fill_opacity=density_fill_opacity,
            showlegend=showlegend,
            visible=True,
        )
        if show_bars:
            traces.append(hist_trace)
        if show_density:
            traces.append(density_trace)
        metrics_list.append(compute_metrics(data, threshold))
    return traces, metrics_list


def plot_hist_density(
    traces: List[Trace],
    threshold: float = 11,
    title: Optional[str] = None,
    threshold_xshift: float = 0,
    threshold_yshift: float = 0,
) -> go.Figure:
    annotations = [
        dict(
            x=threshold + threshold_xshift,
            y=0.8 + threshold_yshift,
            # xref="paper",
            yref="paper",
            yanchor="bottom",
            text=f"<b>Threshold: {threshold}</b>",
            showarrow=False,
        )
    ]
    shapes = [
        dict(
            type="line",
            x0=threshold,
            x1=threshold,
            y0=0,
            y1=1,
            yref="paper",  # refers to the entire plot for the y-dimension
            line=dict(color=Style.biscale["dark_grey"][0](0.6), width=4, dash="dash"),
        )
    ]
    fig = go.Figure(
        data=traces,
        layout=go.Layout(
            bargap=0.2,
            barmode="overlay",
            shapes=shapes,
            annotations=annotations,
        ),
    )
    return fig
